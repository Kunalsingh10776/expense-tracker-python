from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os

FILE_NAME = "expenses.xlsx"

# ---------------- FILE SETUP ----------------
if os.path.exists(FILE_NAME):
    wb = load_workbook(FILE_NAME)
else:
    wb = Workbook()

# Expense Sheet
if "Expenses" not in wb.sheetnames:
    expense_sheet = wb.active
    expense_sheet.title = "Expenses"
    expense_sheet.append(["DATE", "MONTH", "CATEGORY", "ITEM", "AMOUNT"])
else:
    expense_sheet = wb["Expenses"]

# Summary Sheet
if "Summary" not in wb.sheetnames:
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.append(["CATEGORY", "TOTAL"])
else:
    summary_sheet = wb["Summary"]

# Charts Sheet
if "Charts" not in wb.sheetnames:
    chart_sheet = wb.create_sheet("Charts")
else:
    chart_sheet = wb["Charts"]

wb.save(FILE_NAME)

# ---------------- FUNCTIONS ----------------
def add_expense():
    date = input("Enter date (DD/MM/YYYY): ")
    month = date[3:5] + "-" + date[6:]
    category = input("Category: ")
    item = input("Description: ")
    amount = float(input("Amount: "))

    expense_sheet.append([date, month, category, item, amount])
    wb.save(FILE_NAME)
    print("✅ Expense added & saved to Excel")

def view_expenses():
    if expense_sheet.max_row == 1:
        print("❌ No expenses found")
        return

    print("\n📋 ALL EXPENSES")
    for i, row in enumerate(expense_sheet.iter_rows(min_row=2, values_only=True), 1):
        print(f"{i}. {row[0]} | {row[2]} | {row[3]} | ₹{row[4]}")

def total_expenses():
    total = sum(row[4] for row in expense_sheet.iter_rows(min_row=2, values_only=True))
    print(f"\n💰 TOTAL EXPENSE: ₹{total}")

def category_summary():
    summary_sheet.delete_rows(2, summary_sheet.max_row)

    totals = {}
    for row in expense_sheet.iter_rows(min_row=2, values_only=True):
        totals[row[2]] = totals.get(row[2], 0) + row[4]

    for cat, amt in totals.items():
        summary_sheet.append([cat, amt])

    wb.save(FILE_NAME)
    print("📊 Category summary updated")

def monthly_summary():
    monthly = {}
    for row in expense_sheet.iter_rows(min_row=2, values_only=True):
        monthly[row[1]] = monthly.get(row[1], 0) + row[4]

    print("\n📅 MONTHLY SUMMARY")
    for month, amt in monthly.items():
        print(f"{month}: ₹{amt}")

def create_chart():
    chart_sheet.delete_rows(1, chart_sheet.max_row)

    for row in summary_sheet.iter_rows(values_only=True):
        chart_sheet.append(row)

    chart = BarChart()
    chart.title = "Expenses by Category"
    chart.x_axis.title = "Category"
    chart.y_axis.title = "Amount"

    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=chart_sheet.max_row)
    categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=chart_sheet.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart_sheet.add_chart(chart, "E2")

    wb.save(FILE_NAME)
    print("📈 Chart created in Excel")

# ---------------- MENU ----------------
while True:
    print("\n===== EXPENSE TRACKER =====")
    print("1. Add Expense")
    print("2. View Expenses")
    print("3. Total Expense")
    print("4. Category Summary")
    print("5. Monthly Summary")
    print("6. Create Excel Chart")
    print("7. Exit")

    try:
        choice = int(input("Choose option: "))
    except ValueError:
        print("❌ Enter a number only")
        continue

    if choice == 1:
        add_expense()
    elif choice == 2:
        view_expenses()
    elif choice == 3:
        total_expenses()
    elif choice == 4:
        category_summary()
    elif choice == 5:
        monthly_summary()
    elif choice == 6:
        category_summary()
        create_chart()
    elif choice == 7:
        print("👋 Goodbye!")
        break
    else:
        print("❌ Invalid option")
